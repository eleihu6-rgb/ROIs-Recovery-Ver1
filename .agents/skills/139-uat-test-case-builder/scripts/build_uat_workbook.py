#!/usr/bin/env python3
"""Reusable UAT test-case workbook builder for ROIS-AI.

Manifest-driven, N-round layout. Learned from the customer reference files in
docs/UAT/ (Thai Airways "Bush" multi-round per-case format + TCAR summary sheet).

Usage:
    python3 build_uat_workbook.py manifest.json

Manifest schema (JSON):
{
  "output": "docs/UAT/F8 PBS UAT Test Cases - V3.0.xlsx",
  "airline": "F8 (Flair)",
  "round_label": "UAT — 3 Test Rounds",
  "rounds": 3,                       # number of test rounds per case (>=1)
  "ref_date": "2026-08-25",
  "summary_title": "F8 (Flair) PBS — UAT Test Cases — Summary",
  "tabs": [
    {"name": "T2 Admin UI", "title": "T2 — Admin UI  (Live & Scenario)  —  Admin user",
     "subtitle": "...", "area": "Live & Scenario admin (Altair)",
     "sources": ["t2_live.json", "t2_scenario.json"]}
  ]
}

Each source JSON: {"cases":[{"id","feature_area","title","preconditions","steps",
"expected","severity","help_ref"}, ...]}  (a bare list is also accepted).
Cases are grouped under a coloured band per distinct feature_area (in input order).
The first tab of the workbook is always an auto-computed Summary (T1)."""
import json, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

NAVY="1F3864"; BLUE="2E5496"; TEAL="1F6B75"; PLUM="6B3F7A"
LBLUE="D6E0F0"; GREY="F2F2F2"; CALC="EEF3FB"; WHITE="FFFFFF"
ROUND_FILLS=[TEAL, PLUM]     # alternating per round
BORDER=Side(style="thin", color="BFBFBF"); BOX=Border(BORDER,BORDER,BORDER,BORDER)
FIRST=5; LAST=2000
RESULT_LIST='"Pass,Fail,Blocked,Partial Pass,Not Tested"'
PRIO_LIST='"High,Medium,Low"'

def fill(hx): return PatternFill("solid", fgColor=hx)
F_HDR=Font(size=10,bold=True,color=WHITE); F_CELL=Font(size=10,color="000000")
F_BAND=Font(size=10,bold=True,color="1F3864"); F_TITLE=Font(size=15,bold=True,color=WHITE)
F_SUB=Font(size=10,color=WHITE); F_CALC=Font(size=10,bold=True,color="1F3864")
WRAP=Alignment(wrap_text=True,vertical="top")
CTR=Alignment(horizontal="center",vertical="center",wrap_text=True)
LFT=Alignment(horizontal="left",vertical="center",indent=1)

def load_cases(base, name):
    p=os.path.join(base,name)
    if not os.path.exists(p):
        print("WARN missing source:",p); return []
    d=json.load(open(p)); return d.get("cases", d if isinstance(d,list) else [])

def col_layout(rounds):
    """Return (cols, c_latest, c_help, ncol). cols = list of (label,width,kind)."""
    cols=[("Test ID",11,"base"),("Feature",18,"base"),("What to Test",30,"base"),
          ("How to Test (Steps)",52,"base"),("Expected Result",42,"base"),("Priority",9,"base")]
    for k in range(1,rounds+1):
        cols+=[("Result",12,f"r{k}"),("Tester",12,f"r{k}"),("Notes / Screenshot",24,f"r{k}")]
    cols+=[("Latest Result",13,"calc"),("Help Reference",26,"base")]
    ncol=len(cols); c_latest=6+rounds*3+1; c_help=c_latest+1
    return cols, c_latest, c_help, ncol

def round_result_cols(rounds):
    return [7+(k-1)*3 for k in range(1,rounds+1)]   # 1-based col index of each round's Result

def build_test_sheet(ws, tab, cases, rounds):
    cols,c_latest,c_help,ncol=col_layout(rounds)
    for i,(_,w,_) in enumerate(cols,1):
        ws.column_dimensions[get_column_letter(i)].width=w
    last=get_column_letter(ncol)
    ws.merge_cells(f"A1:{last}1"); b=ws["A1"]; b.value=tab["title"]; b.font=F_TITLE
    b.fill=fill(NAVY); b.alignment=LFT; ws.row_dimensions[1].height=30
    ws.merge_cells(f"A2:{last}2"); s=ws["A2"]; s.value=tab.get("subtitle",""); s.font=F_SUB
    s.fill=fill(BLUE); s.alignment=LFT; ws.row_dimensions[2].height=18
    ws.row_dimensions[3].height=16; ws.row_dimensions[4].height=26
    # base + calc headers merged over rows 3:4
    for i,(name,_,kind) in enumerate(cols,1):
        if kind in ("base","calc"):
            L=get_column_letter(i); ws.merge_cells(f"{L}3:{L}4")
            c=ws[f"{L}3"]; c.value=name; c.font=F_HDR
            c.fill=fill(NAVY if kind=="calc" else BLUE); c.alignment=CTR; c.border=BOX
            ws[f"{L}4"].border=BOX
    # round group labels (row 3) + sub-headers (row 4)
    rcols=round_result_cols(rounds)
    for k,c0 in enumerate(rcols,1):
        label=f"Round {k} — Initial test" if k==1 else f"Round {k} — Re-test"
        gf=ROUND_FILLS[(k-1)%len(ROUND_FILLS)]
        L0=get_column_letter(c0); L2=get_column_letter(c0+2)
        ws.merge_cells(f"{L0}3:{L2}3")
        g=ws[f"{L0}3"]; g.value=label; g.font=F_HDR; g.fill=fill(gf); g.alignment=CTR; g.border=BOX
        for j in range(c0,c0+3):
            hc=ws.cell(4,j,cols[j-1][0]); hc.font=F_HDR; hc.fill=fill(gf); hc.alignment=CTR; hc.border=BOX
    # data rows
    r=FIRST; cur=None
    for cs in cases:
        area=cs.get("feature_area","")
        if area!=cur:
            ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=ncol)
            bc=ws.cell(r,1,f"▸ {area}"); bc.font=F_BAND; bc.fill=fill(LBLUE)
            bc.alignment=LFT; bc.border=BOX
            for j in range(2,ncol+1):
                x=ws.cell(r,j); x.border=BOX; x.fill=fill(LBLUE)
            ws.row_dimensions[r].height=17; cur=area; r+=1
        zebra=GREY if r%2==0 else WHITE
        base=[cs.get("id",""),cs.get("feature_area",""),cs.get("title",""),
              cs.get("steps",""),cs.get("expected",""),cs.get("severity","Medium")]
        for i,v in enumerate(base,1):
            c=ws.cell(r,i,v); c.font=F_CELL; c.border=BOX; c.fill=fill(zebra)
            c.alignment=CTR if i in (1,6) else WRAP
        for c0 in rcols:
            for j in range(c0,c0+3):
                c=ws.cell(r,j,""); c.font=F_CELL; c.border=BOX; c.fill=fill(zebra)
                c.alignment=CTR if j==c0 else WRAP
        # latest result = last non-blank round-result, else "Not Tested"
        letters=[get_column_letter(c0) for c0 in rcols]
        expr='"Not Tested"'
        for L in letters:  # build from first→last so last wins (wrap outward)
            expr=f'IF({L}{r}<>"",{L}{r},{expr})'
        c=ws.cell(r,c_latest); c.value="="+expr
        c.font=F_CALC; c.border=BOX; c.fill=fill(CALC); c.alignment=CTR
        h=ws.cell(r,c_help,cs.get("help_ref","")); h.font=F_CELL; h.border=BOX
        h.fill=fill(zebra); h.alignment=WRAP
        r+=1
    end=r-1
    dvR=DataValidation(type="list",formula1=RESULT_LIST,allow_blank=True)
    dvP=DataValidation(type="list",formula1=PRIO_LIST,allow_blank=True)
    ws.add_data_validation(dvR); ws.add_data_validation(dvP)
    if end>=FIRST:
        for c0 in rcols:
            L=get_column_letter(c0); dvR.add(f"{L}{FIRST}:{L}{end}")
        dvP.add(f"F{FIRST}:F{end}")
    ws.freeze_panes="C5"; ws.sheet_view.showGridLines=False
    return end, c_latest

def build_summary(ws, m, tabs_meta, rounds):
    ws.sheet_view.showGridLines=False
    SUM=[("Tab",16),("Feature Area",30),("Total",9),("Passed",9),("Failed",9),
         ("Blocked",9),("Partial",9),("Not Tested",11),("Ready for Re-test",15),
         ("Pass Rate %",11),("Tested %",10),("Owner",13),("Last Update",13)]
    for i,(_,w) in enumerate(SUM,1): ws.column_dimensions[get_column_letter(i)].width=w
    lc=get_column_letter(len(SUM))
    ws.merge_cells(f"A1:{lc}1"); t=ws["A1"]
    t.value=f'{m.get("summary_title","UAT Test Cases — Summary")}  ({rounds} rounds)'
    t.font=Font(size=16,bold=True,color=WHITE); t.fill=fill(NAVY); t.alignment=LFT; ws.row_dimensions[1].height=34
    ws.merge_cells(f"A2:{lc}2"); st=ws["A2"]
    st.value=(f'{m.get("round_label","UAT")}   •   Airline: {m.get("airline","")}   •   '
              f'Totals use each case’s Latest Result (newest round filled)   •   {m.get("ref_date","")}')
    st.font=F_SUB; st.fill=fill(BLUE); st.alignment=LFT; ws.row_dimensions[2].height=18
    ws.row_dimensions[3].height=6
    for i,(nm,_) in enumerate(SUM,1):
        c=ws.cell(4,i,nm); c.font=F_HDR; c.fill=fill(BLUE); c.alignment=CTR; c.border=BOX
    ws.row_dimensions[4].height=26
    row=5
    for name,area,c_latest in tabs_meta:
        Llat=get_column_letter(c_latest); q=f"'{name}'"
        rid=f"{q}!A{FIRST}:A{LAST}"; rlat=f"{q}!{Llat}{FIRST}:{Llat}{LAST}"
        ws.cell(row,1,name); ws.cell(row,2,area)
        ws.cell(row,3).value=f'=COUNTIF({rid},"T?-*")'
        ws.cell(row,4).value=f'=COUNTIF({rlat},"Pass")'
        ws.cell(row,5).value=f'=COUNTIF({rlat},"Fail")'
        ws.cell(row,6).value=f'=COUNTIF({rlat},"Blocked")'
        ws.cell(row,7).value=f'=COUNTIF({rlat},"Partial Pass")'
        ws.cell(row,8).value=f'=C{row}-D{row}-E{row}-F{row}-G{row}'
        ws.cell(row,9).value=f'=E{row}+F{row}+G{row}'
        ws.cell(row,10).value=f'=IF(C{row}=0,"-",D{row}/C{row})'
        ws.cell(row,11).value=f'=IF(C{row}=0,"-",(C{row}-H{row})/C{row})'
        row+=1
    tr=row
    ws.cell(tr,1,"TOTAL"); ws.cell(tr,2,"All UAT scope")
    for col in range(3,10):
        L=get_column_letter(col); ws.cell(tr,col).value=f"=SUM({L}5:{L}{tr-1})"
    ws.cell(tr,10).value=f'=IF(C{tr}=0,"-",D{tr}/C{tr})'
    ws.cell(tr,11).value=f'=IF(C{tr}=0,"-",(C{tr}-H{tr})/C{tr})'
    for r in range(5,tr+1):
        tot=(r==tr)
        for c in range(1,len(SUM)+1):
            cell=ws.cell(r,c); cell.border=BOX
            cell.font=Font(size=10,bold=tot,color="1F3864" if tot else "000000")
            cell.fill=fill(LBLUE if tot else (GREY if r%2==0 else WHITE))
            cell.alignment=CTR if c>=3 else LFT
            if c in (10,11): cell.number_format="0%"
        ws.row_dimensions[r].height=20
    notes=[
        f"How to use ({rounds}-round UAT):",
        "1. Round 1 — run every case, fill Round 1 → Result (Pass / Fail / Blocked / Partial Pass), Tester, Notes/Screenshot.",
        "2. Any case that is not Pass is re-tested in the next round after the fix; record it under that round.",
        "3. 'Latest Result' auto-shows the newest round you filled — that is what the Summary counts (fail R1 then pass R2 = Passed).",
        "4. 'Ready for Re-test' = cases whose latest result is Fail / Blocked / Partial — the queue for the next round.",
    ]
    for i,txt in enumerate(notes):
        c=ws.cell(tr+2+i,1,txt); c.font=Font(size=10,bold=(i==0),color="1F3864" if i==0 else "404040")
    # Change Log (after How-to-use) — data-driven from manifest["changelog"], newest version first
    rc=tr+2+len(notes)+1
    cl=m.get("changelog",[])
    if cl:
        h=ws.cell(rc,1,"Change Log:"); h.font=Font(size=11,bold=True,color="1F3864"); rc+=1
        for entry in cl:
            tag=f'{entry.get("version","")} — {entry.get("date","")}'
            if entry.get("current"): tag+="   (current)"
            hc=ws.cell(rc,1,tag); hc.font=Font(size=10,bold=True,color="2E5496"); rc+=1
            for it in entry.get("items",[]):
                ic=ws.cell(rc,1,f"     •  {it}"); ic.font=Font(size=10,color="404040"); rc+=1
            rc+=1   # blank line between versions
    ws.freeze_panes="A5"

def main():
    if len(sys.argv)<2:
        print("usage: build_uat_workbook.py manifest.json"); sys.exit(1)
    mpath=os.path.abspath(sys.argv[1]); base=os.path.dirname(mpath)
    m=json.load(open(mpath))
    rounds=int(m.get("rounds",3)); rounds=max(1,rounds)
    wb=Workbook(); ws1=wb.active; ws1.title="T1 Summary"
    tabs_meta=[]
    for tab in m["tabs"]:
        cases=[]
        for src in tab["sources"]:
            cases+=load_cases(base, src)
        ws=wb.create_sheet(tab["name"])
        _, c_latest=build_test_sheet(ws, tab, cases, rounds)
        tabs_meta.append((tab["name"], tab.get("area",""), c_latest))
        print(f"  {tab['name']}: {len(cases)} cases")
    build_summary(ws1, m, tabs_meta, rounds)
    out=m["output"]
    if not os.path.isabs(out): out=os.path.join(base, out)
    wb.save(out); print("SAVED:", out)

if __name__=="__main__":
    main()
