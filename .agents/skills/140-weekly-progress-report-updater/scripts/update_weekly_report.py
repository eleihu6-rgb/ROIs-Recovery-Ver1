#!/usr/bin/env python3
"""Add a new weekly tab to the F8 project status workbook (one tab per week).

Copies SOURCE_TAB (the most recent week's tab) to NEW_TAB, leaving prior weeks untouched, then
on the NEW tab only: rolls Report Date / Report Period to this Tuesday and rewrites the two
High-Level Review narrative cells (D8 = items 1-3, D9 = items 4-6) as rich text so that
NEW/changed items this week render in blue and carried-over items stay black.

Usage:
    python3 update_weekly_report.py "/path/to/F8-CMS-CrewWeeklyReport <DD><Mon><YYYY>.xlsx"

Edit the CONFIG block (SOURCE_TAB, NEW_TAB, dates) and the D8_ITEMS / D9_ITEMS lists each
week, then run.
Each item is a tuple: (text, is_new, is_bold)
    is_new  -> True  = this week's new/changed line -> BLUE
             False   = carried over / unchanged      -> BLACK  (reserve existing content)
    is_bold -> True for section titles ("2. Rule Engine - on track", ...)
Verify by re-reading with rich_text=True (this script prints a colour receipt at the end).
"""
import sys
import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Font

# ---------------------------------------------------------------- CONFIG (edit weekly)
SOURCE_TAB    = "11Aug26 Rpt"                   # most recent existing tab (copy source)
NEW_TAB       = "25Aug26 Rpt"                   # this week's new tab <DD><Mon><YY> Rpt
REPORT_DATE   = "2026/08/25"                    # this Tuesday (meeting day)
REPORT_PERIOD = "2026/08/18 - 2026/08/25"       # last Tuesday - this Tuesday

BLUE  = "FF0070C0"
BLACK = "FF000000"
FONT  = "Arial"
SIZE  = 14

# D8 = High-Level Review items 1 (UAT) / 2 (Rule Engine) / 3 (Data & Integration)
D8_ITEMS = [
    ("1. UAT Preparation\n", False, True),
    ("     • Flair to prepare UAT test cases.\n", False, False),
    ("2. Rule Engine — on track\n", False, True),
    ("     (a) Rule-by-rule verification in progress.\n", False, False),
    ("     (d) “Crew members not fly together” rule expanded to cover crew across "
     "different bases and different divisions.\n", True, False),
    ("3. Data & Integration — on track  (main focus this week)\n", False, True),
    ("     (b) Input interface — RosterGround: DHD overlapping tasks filtered out on import; "
     "assignments already carrying a pairing (PairingId > 0) are not recreated.\n", True, False),
    ("     (d) SSO — login interface developed; UAT environment configuration pending.", True, False),
]

# D9 = High-Level Review items 4 (AI Core) / 5 (Crew Portal UI) / 6 (Admin Portal UI)
D9_ITEMS = [
    ("4. AI Core — on track\n", False, True),
    ("     (b) Coverage & metrics — continuous optimization (in progress) (PBS-458).\n", True, False),
    ("5. Crew Portal UI — on track\n", False, True),
    ("     (e) Reserve page retired — its data is now shown in the left mini-calendar.\n", True, False),
    ("6. Admin Portal UI — on track\n", False, True),
    ("     (a) Gantt menu permission control — in progress; fixing remaining bugs.", True, False),
]
# ----------------------------------------------------------------------------------------


def _blk(text, is_new, is_bold):
    return TextBlock(InlineFont(rFont=FONT, sz=SIZE, b=is_bold,
                                color=(BLUE if is_new else BLACK)), text)


def _rich(items):
    return CellRichText([_blk(t, n, b) for (t, n, b) in items])


def main(path):
    wb = openpyxl.load_workbook(path)
    if SOURCE_TAB not in wb.sheetnames:
        sys.exit(f"source tab {SOURCE_TAB!r} not found; tabs = {wb.sheetnames}")
    if NEW_TAB in wb.sheetnames:
        sys.exit(f"tab {NEW_TAB!r} already exists — refusing to overwrite")

    ws = wb.copy_worksheet(wb[SOURCE_TAB])   # prior tabs stay untouched
    ws.title = NEW_TAB
    wb.active = wb.sheetnames.index(NEW_TAB)

    for ref, val in (("C2", REPORT_DATE), ("C3", REPORT_PERIOD)):
        c = ws[ref]
        c.value = val
        c.font = Font(name=c.font.name, size=c.font.size, bold=c.font.bold,
                      italic=c.font.italic, color=BLUE)  # changed this week -> blue

    for ref, items, height in (("D8", D8_ITEMS, 300), ("D9", D9_ITEMS, 470)):
        ws[ref] = _rich(items)
        ws[ref].alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    ws.row_dimensions[8].height = 300
    ws.row_dimensions[9].height = 470

    wb.save(path)

    # ---- receipt (§No-Illusion) ----
    wb2 = openpyxl.load_workbook(path, rich_text=True)
    ws2 = wb2[NEW_TAB]
    print("SAVED:", path)
    print("SHEETS:", wb2.sheetnames)
    print(f"NEW TAB {NEW_TAB}: C2={ws2['C2'].value} | C3={ws2['C3'].value} | "
          f"merged={len(list(ws2.merged_cells.ranges))}")
    for ref in ("D8", "D9"):
        blue = sum(1 for p in ws2[ref].value
                   if getattr(p, "font", None) and p.font.color and p.font.color.rgb == BLUE)
        total = len(ws2[ref].value)
        print(f"  {ref}: {total} segments, {blue} blue (new this week)")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        sys.exit("usage: python3 update_weekly_report.py <workbook.xlsx>")
    main(sys.argv[1])
